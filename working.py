#!/usr/bin/python3
#m7while.py With the objective of explain the while used By Luis Najera
#For www.codewithcharlie.com
#Begins Import stage
import sys
import termcolor
import datetime
import pprint
import tweepy
import getpass
import passwd
from termcolor import colored, cprint
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
#Begins Function Definition

def cRed(string):
    string = colored(string,"red")
    return string

def cBlue(string):
    string = colored(string,"blue")
    return string

def cGreen(string):
    string = colored(string,"green")
    return string

def cYellow(string):
    string = colored(string,"yellow")
    return string

def bBlue(string,colortoblink):
    string = colored(string, colortoblink, attrs=['reverse', 'blink'])
    return string


def uMauth(aName):
    print()
    print(cBlue("Buenos dias "+aName+" ¿Con que te puedo ayudar?"))
    print()
    print("Para Reportar Horario de Entrada Introduce 0")
    print()
    print("Para Reportar Horario de Salida Introduce 1")
    print()
    print("Para Reportar Falta Introduce 2 ")
    print()
    print("Para Reportar Gasto Introduce 3")
    print()
    print("Para Reporte de ganancias por empleada Introduce 4 ")
    print()

def opc0(confirmacion):

    while(confirmacion == "NO"):
        print()
        fName=input("Introduce el Nombre de la Trabajadora ")
        print()
        lName=input("Introduce el Apellido de la Trabajadora ")
        print()
        horario=input("Introduce la hora de la computadora en el formato HH:MM ")
        print()
        print("¿La informacion es correcta? Nombre:",fName,"Apellido:",lName,"Y la hora es",horario)
        print()
        time = datetime.datetime.now().time()
        time=time.isoformat()
        confirmacion = input("¿Si o No? ")
        confirmacion = confirmacion.upper()
        maEnviar = "Reporte de entrada. Nombre de Trabajadora "+fName+" "+lName+" Horario Reportado "+horario+" Horario Real "+time
    return(maEnviar+" Generado por "+aName)

def opc1(confirmacion):
    while(confirmacion == "NO"):
        print()
        fName=input("Introduce el Nombre de la Trabajadora ")
        print()
        lName=input("Introduce el Apellido de la Trabajadora ")
        print()
        horario=input("Introduce la hora de la computadora en el formato HH:MM ")
        print()
        print("¿La informacion es correcta? Nombre:",fName,"Apellido:",lName,"Y la hora es",horario)
        print()
        time = datetime.datetime.now().time()
        time=time.isoformat()
        confirmacion = input("¿Si o No? ")
        confirmacion = confirmacion.upper()
        maEnviar = "Reporte de salida. Nombre de Trabajadora "+fName+" "+lName+" Horario Reportado "+horario+" Horario Real "+time
    return(maEnviar+" Generado por "+aName)

def opc2(confirmacion):
    while(confirmacion == "NO"):
        print()
        fName=input("Introduce el Nombre de la Trabajadora ")
        print()
        lName=input("Introduce el Apellido de la Trabajadora ")
        print()
        motivo=input("Introduce el motivo de la falta si se especifico ")
        print()
        text="Se Mando Foto del Justificante? "
        text=cRed(text)
        justificante=input(text)
        print()
        time = datetime.datetime.now().time()
        time=time.isoformat()
        maEnviar = "Reporte de Falta. Nombre de Trabajadora "+fName+" "+lName+" Motivo Reportado "+motivo+" "+justificante+" Se mando Justificante "+"Horario de reporte "+time
        maEnviar=cBlue(maEnviar)
        confirmacion=input("Los Datos Son Correctos? ")
        confirmacion = confirmacion.upper()
    return(maEnviar+" Generado por "+aName)

def opc3(confirmacion): #Gastos
    while(confirmacion == "NO"):
        print()
        text   = "Introduce la razon del gasto "
        nGasto = input(cGreen(text))
        print()
        text   = "Introduce la Cantidad del gasto en el formato PPP.CC "
        pGasto = float(input(cBlue(text)))
        print()
        print("Confirmar, La razon del gasto fue "+nGasto+" por la cantidad de $",pGasto)
        text="¿Si o No? "
        confirmacion = input(cYellow(text))
        print()
        confirmacion = confirmacion.upper()
        pGasto = -pGasto
    time = datetime.datetime.now().time()
    time=time.isoformat()
    maEnviar = "Se Reporto un gasto con razon de "+nGasto+" por ",pGasto," reportado por "+aName+" a las "+time
    ws1['A'+ist]='Gasto'
    ws1['B'+ist]=aName
    ws1['C'+ist]=time
    ws1['D'+ist]=aName
    ws1["E"+ist]=nGasto
    ws1['F'+ist]=pGasto
    ws1["G2"]=i+1

    return(maEnviar)


def opc4(confirmacion): #Pagos
        print()
        text = "Introduce el nombre de la Empleada "
        fName=input(cRed(text))
        print()
        text="Introduce el Apellido de la Empleada "
        lName=input(cGreen(text))
        print()
        text=fName+" Recibe Sueldo? "
        sueldo = input(cBlue(text))
        print()
        sueldo = sueldo.upper()
        if(sueldo == "SI"):

            while(confirmacion == "NO"):
                print()
                text= "Introduce el sueldo Semanal en formato PPP.cc "
                pago=input(cGreen(text))
                print()
                text="Introduce el porcentaje de comision que le corresponde sin el % "
                comision=input(cBlue(text))
                print()
                text="Cual fue su venta total en la semana? (formato PPP.cc) "
                vtotal = input(cBlue(text))
                print()
                text = "El sueldo semanal de "+fName+" es de $"+pago+", su venta semanal fue de $"+vtotal+" el porcentaje de comision es %"+comision+"?"
                print(cYellow(text))
                confirmacion = input("La informacion es correcta ¿Si o No? ")
                confirmacion = confirmacion.upper()

                fSueldo = float(pago)
                fVenta  = float(vtotal)
                fPct    = float(comision)
                vCom = (fPct/100)*fVenta
                pagoFin = (fSueldo)+(fVenta)*(fPct/100)
                stPag = str(pagoFin)
                sCom = str(vCom)
                time = datetime.datetime.now().time()
                time=time.isoformat()
                maEnviar="Se pago a "+fName+" "+lName+" la cantidad de $"+stPag+" De los cuales $"+pago+" son de su sueldo y $"+sCom+" son de Comision Reportado por "+aName+" A las "+time
                api.send_direct_message(screen_name='lcnajerao', text=maEnviar)
                print("El total a pagar es de $",pagoFin )
                vCom = -vCom
                fSueldo = -fSueldo

                i = sheet_ranges["G2"].value
                ist = str(i)

                ws1['A'+ist]='Venta Semanal'
                ws1['B'+ist]=aName
                ws1['C'+ist]=time
                ws1['D'+ist]=fName+" "+lName
                ws1["E"+ist]="Venta"
                ws1['F'+ist]=fVenta
                ws1["G2"]=i+1
                i = sheet_ranges["G2"].value
                ist = str(i)

                ws1['A'+ist]='Sueldo'
                ws1['B'+ist]=aName
                ws1['C'+ist]=time
                ws1['D'+ist]=fName+" "+lName
                ws1["E"+ist]="Sueldo"
                ws1['F'+ist]=fSueldo
                ws1["G2"]=i+1
                i = sheet_ranges["G2"].value
                ist = str(i)

                ws1['A'+ist]='Comision'
                ws1['B'+ist]=aName
                ws1['C'+ist]=time
                ws1['D'+ist]=fName+" "+lName
                ws1["E"+ist]="Sueldo"
                ws1['F'+ist]=vCom
                ws1["G2"]=i+1
                i = sheet_ranges["G2"].value
                ist = str(i)


        elif(sueldo == "NO"):

            while(confirmacion == "NO"):
                print()
                text="Introduce el porcentaje de comision que le corresponde sin el % "
                comision=input(cBlue(text))
                print()
                text="Cual fue su venta total en la semana? (formato PPP.cc) "
                vtotal = input(cBlue(text))
                print()
                text = "La venta semanal de "+fName+" es de $"+vtotal+" y el porcentaje de comision es %"+comision+"?"
                print(cYellow(text))
                print()
                confirmacion = input("La informacion es correcta ¿Si o No? ")
                print()
                confirmacion = confirmacion.upper()
                fVenta  = float(vtotal)
                fPct    = float(comision)
                pagoFin = fVenta*(fPct/100)
                stPag = str(pagoFin)
                sCom = str(fPct)
                time = datetime.datetime.now().time()
                time=time.isoformat()
                maEnviar="Se pago a "+fName+" "+lName+" la cantidad de $"+stPag+" De los cuales $"+sCom+" son de Comision Reportado por "+aName+" A las "+time
                api.send_direct_message(screen_name='lcnajerao', text=maEnviar)

                i = sheet_ranges["G2"].value
                ist = str(i)
                pagoFin = -pagoFin

                ws1['A'+ist]='Venta Semanal'
                ws1['B'+ist]=aName
                ws1['C'+ist]=time
                ws1['D'+ist]=fName+" "+lName
                ws1["E"+ist]="Venta"
                ws1['F'+ist]=fVenta
                ws1["G2"]=i+1
                i = sheet_ranges["G2"].value
                ist = str(i)

                ws1['A'+ist]='Comision'
                ws1['B'+ist]=aName
                ws1['C'+ist]=time
                ws1['D'+ist]=fName+" "+lName
                ws1["E"+ist]="Comision"
                ws1['F'+ist]=pagoFin
                ws1["G2"]=i+1
                i = sheet_ranges["G2"].value
                ist = str(i)

def pyxl():
    wb = load_workbook(filename = "test.xlsx")
    sheet_ranges = wb["Ingresos Y Egresos"]
    i = sheet_ranges["G2"].value
    print(i)
    ist = str(i)
    dest_filename = "test.xlsx"
    ws1 = wb.active
    ws1.title = "Ingresos Y Egresos"




#Begins Variable Definition
cKey = passwd.cKey
cSec = passwd.cSec
cWeb = "z"
cCod = "p"
cAut = "s"
pBli = "x"
time = "10:10"
fName = "Luis"
lName = "Najera"
motivo = "x"
password = "luisrifa"
passtrial = "x"
opcion = 9
confirmacion = "NO"
horario = "hh:mm"
maEnviar="X"
justificante = "x"
aName="YO"
loop="si"
nGasto="Line2U"
pGasto=0
sueldo = "x"
pago = "0"
fPago = 0.00
fPct = 0.00
fSueldo = 0.00
fVenta = 0.00
pagoFin = 0.00
comision = "0"
vtotal="0"
vCom=0.0
time = 'x'
password = 'x'
i = 0
ist = "x"
#Begins User Interface
auth = tweepy.OAuthHandler(passwd.cKey, passwd.cSec)
auth.set_access_token(passwd.aTok,passwd.aSec)
api = tweepy.API(auth)
#cWeb = auth.get_authorization_url()
#text = cYellow(cWeb)
#text2 = "Ve a la Pagina de internet:"
#text2 = cBlue(text2)
#print(text2,text)
#text = colored('Copia el codigo de Seguridad aqui --> ', 'red')
#cCod = input(text)
#cAut = auth.get_access_token(cCod)

passtrial = getpass.getpass("Introduce la contraseña del administrador ")
wb = load_workbook(filename = "test.xlsx")
sheet_ranges = wb["Ingresos Y Egresos"]
i = sheet_ranges["G2"].value
print(i)
ist = str(i)
dest_filename = "test.xlsx"
ws1 = wb.active
ws1.title = "Ingresos Y Egresos"
while(loop=="si"):
    if(passtrial == passwd.rosyeli):

        aName = "Eli Garcia "
        uMauth(aName)
        intop=cBlue("Introduce La Opcion Deseada ")
        opcion = int(input(intop))

        if(opcion == 0):
            text=opc0(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 1):
            text=opc1(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 2):
            text=opc2(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 3):
            text=opc3(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 4):

            text=opc4(confirmacion)

    elif(passtrial == passwd.carlitos):

        aName = "Luis Carlos Najera "
        uMauth(aName)
        intop=cBlue("Introduce La Opcion Deseada ")
        opcion = int(input(intop))

        if(opcion == 0):
            text=opc0(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 1):
            text=opc1(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 2):
            text=opc2(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 3):
            text=opc3(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 4):

            text=opc4(confirmacion)


    elif(passtrial == passwd.maryeli):
        aName = "Maria Elizabeth "
        uMauth(aName)
        intop=cBlue("Introduce La Opcion Deseada ")
        opcion = int(input(intop))

        if(opcion == 0):
            text=opc0(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 1):
            text=opc1(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 2):
            text=opc2(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 3):
            text=opc3(confirmacion)
            api.send_direct_message(screen_name='lcnajerao', text=text)

        elif(opcion == 4):

            text=opc4(confirmacion)

    else:
        print("Error")
    wb.save(filename = 'test.xlsx')
    text="Deseas realizar otra operacion? (Si o No) "
    loop=input(cGreen(text))
